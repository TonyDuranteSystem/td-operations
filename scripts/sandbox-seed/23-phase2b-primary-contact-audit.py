"""
Phase 2 area 2 — account_contacts.is_primary backfill AUDIT (read-only).

For each active account (status NOT IN Cancelled/Closed), propose which
contact_id should be is_primary=true:
  - 1 contact on the account → that contact (HIGH confidence)
  - 2+ contacts → the signer of the formation/onboarding contract for that account
                  (HIGH if exactly one signer; MEDIUM if multiple; LOW if no
                  formation/onboarding contract found → needs Antonio)

Output: ~/Downloads/Phase2-PrimaryContact-Proposal.xlsx
"""
import os, psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_URL = None
with open(os.path.expanduser('~/Developer/td-operations/.env.sandbox')) as f:
    for line in f:
        if line.startswith('SUPABASE_DB_URL='):
            DB_URL = line.split('=', 1)[1].strip().strip('"')
            break
assert DB_URL

conn = psycopg2.connect(DB_URL, sslmode='require')
cur = conn.cursor()

# Active accounts with their contact links + contact names
cur.execute("""
    SELECT a.id AS account_id, a.company_name, a.entity_type::text AS entity_type, a.status::text AS status,
           ac.contact_id, c.full_name, c.email, ac.role, ac.is_primary
    FROM accounts a
    JOIN account_contacts ac ON ac.account_id = a.id
    JOIN contacts c ON c.id = ac.contact_id
    WHERE a.status NOT IN ('Cancelled','Closed')
    ORDER BY a.company_name
""")
rows = cur.fetchall()

# Group by account
from collections import defaultdict
by_acct = defaultdict(list)
accounts_meta = {}
for r in rows:
    aid, name, etype, status, cid, cname, cemail, role, is_primary = r
    by_acct[aid].append({'contact_id': cid, 'name': cname, 'email': cemail, 'role': role, 'is_primary': is_primary})
    accounts_meta[aid] = {'name': name, 'entity_type': etype, 'status': status}

# For multi-contact accounts, look up the signer of formation/onboarding contract
multi_ids = [aid for aid, cs in by_acct.items() if len(cs) > 1]
signer_by_account = {}
if multi_ids:
    cur.execute("""
        SELECT ct.account_id, ct.contact_id, c.full_name, o.contract_type
        FROM contracts ct
        LEFT JOIN offers o ON o.token = ct.offer_token
        LEFT JOIN contacts c ON c.id = ct.contact_id
        WHERE ct.account_id = ANY(%s::uuid[])
          AND o.contract_type IN ('formation','onboarding')
          AND ct.contact_id IS NOT NULL
        ORDER BY ct.signed_at
    """, ([str(x) for x in multi_ids],))
    for aid, cid, cname, ctype in cur.fetchall():
        signer_by_account.setdefault(aid, []).append({'contact_id': cid, 'name': cname, 'contract_type': ctype})

out_rows = []
for aid, contacts_list in by_acct.items():
    meta = accounts_meta[aid]
    current_primary = [c for c in contacts_list if c['is_primary']]
    proposed_contact_id = None
    proposed_name = None
    confidence = 'LOW'
    reason = ''

    if len(contacts_list) == 1:
        c = contacts_list[0]
        proposed_contact_id = c['contact_id']
        proposed_name = c['name']
        confidence = 'HIGH'
        reason = 'only contact on this account'
    else:
        signers = signer_by_account.get(aid, [])
        unique_signers = {s['contact_id']: s for s in signers}
        if len(unique_signers) == 1:
            s = next(iter(unique_signers.values()))
            proposed_contact_id = s['contact_id']
            proposed_name = s['name']
            confidence = 'HIGH'
            reason = f'only signer of {s["contract_type"]} contract'
        elif len(unique_signers) > 1:
            # multiple signers, pick the most recent formation/onboarding? surface for review
            proposed_contact_id = None
            proposed_name = ' / '.join(s['name'] for s in unique_signers.values())
            confidence = 'LOW'
            reason = f'{len(unique_signers)} different signers on formation/onboarding contracts — needs Antonio'
        else:
            proposed_contact_id = None
            proposed_name = '(none)'
            confidence = 'LOW'
            reason = 'multiple contacts but NO formation/onboarding contract found — needs Antonio'

    existing_names = ' / '.join(c['name'] for c in contacts_list)
    current_primary_name = current_primary[0]['name'] if current_primary else '(none)'

    if confidence == 'LOW':
        action = 'NEEDS REVIEW'
    elif current_primary and current_primary[0]['contact_id'] == proposed_contact_id:
        action = 'ALREADY CORRECT'
    else:
        action = 'AUTO-APPLY'

    out_rows.append({
        'Action': action,
        'Company': meta['name'],
        'Entity Type': (meta['entity_type'] or '').replace('Single Member LLC', 'SMLLC').replace('Multi Member LLC', 'MMLLC'),
        'Total Contacts': len(contacts_list),
        'All Contacts': existing_names,
        'Current Primary': current_primary_name,
        'Proposed Primary': proposed_name or '(none)',
        'Confidence': confidence,
        'Reason': reason,
    })

# Sort: NEEDS REVIEW first, then AUTO-APPLY, then ALREADY CORRECT
priority = {'NEEDS REVIEW': 0, 'AUTO-APPLY': 1, 'ALREADY CORRECT': 2}
out_rows.sort(key=lambda r: (priority.get(r['Action'], 99), r['Company']))

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = 'Primary Contact Proposal'

headers = list(out_rows[0].keys())
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', start_color='305496')
review_fill = PatternFill('solid', start_color='FFF2CC')
apply_fill = PatternFill('solid', start_color='E2EFDA')
correct_fill = PatternFill('solid', start_color='F2F2F2')
thin = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for i, r in enumerate(out_rows, start=2):
    fill = review_fill if r['Action']=='NEEDS REVIEW' else apply_fill if r['Action']=='AUTO-APPLY' else correct_fill
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=i, column=c, value=r[h])
        cell.fill = fill; cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

widths = {'Action': 18, 'Company': 36, 'Entity Type': 14, 'Total Contacts': 12,
          'All Contacts': 44, 'Current Primary': 24, 'Proposed Primary': 26,
          'Confidence': 12, 'Reason': 52}
for col_idx, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 18)
ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'

# Summary sheet
ws2 = wb.create_sheet('Summary')
ws2['A1'] = 'Total active accounts'; ws2['B1'] = len(out_rows)
ws2['A2'] = 'AUTO-APPLY'; ws2['B2'] = sum(1 for r in out_rows if r['Action']=='AUTO-APPLY')
ws2['A3'] = 'NEEDS REVIEW'; ws2['B3'] = sum(1 for r in out_rows if r['Action']=='NEEDS REVIEW')
ws2['A4'] = 'ALREADY CORRECT (no change)'; ws2['B4'] = sum(1 for r in out_rows if r['Action']=='ALREADY CORRECT')
ws2.column_dimensions['A'].width = 36; ws2.column_dimensions['B'].width = 10
for row in (1,2,3,4):
    ws2[f'A{row}'].font = Font(bold=True)

out = os.path.expanduser('~/Downloads/Phase2-PrimaryContact-Proposal.xlsx')
wb.save(out)
print(f'wrote {out}')
print(f'rows: {len(out_rows)}')
counts = {}
for r in out_rows: counts[r['Action']] = counts.get(r['Action'], 0) + 1
print(f'action breakdown: {counts}')
cur.close(); conn.close()
