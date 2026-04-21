"""
Convert Phase 2 area 1 audit into a human-readable Excel workbook.

- Pulls contract → offer → contact → account in one sandbox query
- Resolves UUIDs into company names / contact names
- Labels confidence in plain English
- Groups rows: ACTION NEEDED / AUTO-APPLY / SKIP (duplicates)
- Color codes the status column
"""
import os, psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_URL = None
with open(os.path.expanduser('~/Developer/td-operations/.env.sandbox')) as f:
    for line in f:
        if line.startswith('SUPABASE_DB_URL='):
            DB_URL = line.split('=', 1)[1].strip().strip('"')
            break
assert DB_URL, 'SUPABASE_DB_URL not found'

conn = psycopg2.connect(DB_URL, sslmode='require')
cur = conn.cursor()

cur.execute("""
    SELECT
      ct.id, ct.offer_token, ct.client_name, ct.client_email,
      ct.llc_type, ct.annual_fee, ct.signed_at, ct.status,
      o.contract_type, o.account_id, o.lead_id, o.services,
      l.converted_to_contact_id,
      a.company_name AS offer_company,
      a.formation_date AS account_formation_date
    FROM contracts ct
    LEFT JOIN offers o ON o.token = ct.offer_token
    LEFT JOIN leads  l ON l.id = o.lead_id
    LEFT JOIN accounts a ON a.id = o.account_id
    ORDER BY ct.signed_at
""")
contracts = cur.fetchall()
cols = [d[0] for d in cur.description]

# identify duplicate-token rows (area 6 material)
_tok_counts = {}
for r in contracts:
    tok = dict(zip(cols, r))['offer_token']
    _tok_counts[tok] = _tok_counts.get(tok, 0) + 1
DUPLICATE_TOKENS = {tok for tok, n in _tok_counts.items() if n > 1}

# lookup helpers
def dict_of(row):
    return dict(zip(cols, row))

# Resolve contacts by email
emails = sorted({dict_of(r)['client_email'] for r in contracts if dict_of(r)['client_email']})
cur.execute(
    "SELECT id, email, full_name FROM contacts WHERE email = ANY(%s)",
    (emails,)
)
contacts_by_email = {}
for cid, email, name in cur.fetchall():
    contacts_by_email.setdefault((email or '').lower(), []).append((cid, name))

# Resolve contact names by id (for converted_to_contact_id)
converted_ids = [dict_of(r)['converted_to_contact_id'] for r in contracts if dict_of(r)['converted_to_contact_id']]
contact_name_by_id = {}
if converted_ids:
    cur.execute("SELECT id, full_name FROM contacts WHERE id = ANY(%s::uuid[])", ([str(x) for x in converted_ids],))
    for cid, name in cur.fetchall():
        contact_name_by_id[cid] = name

# Load candidate payments per account
offer_account_ids = sorted({dict_of(r)['account_id'] for r in contracts if dict_of(r)['account_id']})
payments_by_acct = {}
if offer_account_ids:
    cur.execute("""
        SELECT id, account_id, amount, invoice_number, status::text, paid_date, issue_date, due_date, description
        FROM payments WHERE account_id = ANY(%s::uuid[])
    """, ([str(x) for x in offer_account_ids],))
    for pid, acct, amt, inv, st, pd, id_, dd, desc in cur.fetchall():
        payments_by_acct.setdefault(acct, []).append({
            'id': pid, 'amount': amt, 'invoice_number': inv, 'status': st,
            'paid_date': pd, 'issue_date': id_, 'due_date': dd, 'description': desc
        })

PAYMENT_TYPE_MAP = {
    'formation': 'Setup Fee',
    'onboarding': 'Setup Fee',
    'renewal': 'Installments',
    'tax_return': 'One-Time Service',
    'itin': 'One-Time Service',
    'custom': 'Custom',
}

def human_contract_type(t):
    return {
        'formation': 'Formation (new LLC)',
        'onboarding': 'Onboarding (existing LLC)',
        'renewal': 'Annual Renewal',
        'tax_return': 'Standalone Tax Return',
        'itin': 'Standalone ITIN',
        'custom': 'Custom',
    }.get((t or '').lower(), t or '(unknown)')

rows_out = []
for rec in contracts:
    r = dict_of(rec)

    # Contact resolution
    contact_name = None
    contact_note = ''
    contact_conf = 'LOW'
    if r['converted_to_contact_id']:
        contact_name = contact_name_by_id.get(r['converted_to_contact_id'], '(contact id ' + str(r['converted_to_contact_id'])[:8] + ')')
        contact_note = 'via signed-offer lead'
        contact_conf = 'HIGH'
    elif r['client_email']:
        matches = contacts_by_email.get((r['client_email'] or '').lower(), [])
        if len(matches) == 1:
            contact_name = matches[0][1]
            contact_note = 'matched by email'
            contact_conf = 'HIGH'
        elif len(matches) > 1:
            contact_name = ' / '.join(m[1] for m in matches)
            contact_note = f'ambiguous — {len(matches)} contacts share this email'
            contact_conf = 'LOW'
        else:
            contact_note = 'no contact record found for this email'
            contact_conf = 'LOW'
    else:
        contact_note = 'no email, no lead'
        contact_conf = 'LOW'

    # Account resolution
    company = r['offer_company'] or ('(contact-only — no company yet)')
    account_conf = 'HIGH' if r['account_id'] else 'N/A (pre-account)'

    # Payment type
    ct = (r['contract_type'] or '').lower()
    payment_type = PAYMENT_TYPE_MAP.get(ct, 'Custom')

    # Invoice match
    invoice_label = ''
    invoice_note = ''
    if r['account_id'] and r['account_id'] in payments_by_acct:
        fee = None
        try: fee = float(r['annual_fee']) if r['annual_fee'] else None
        except: fee = None
        signed = r['signed_at']
        best = None
        for p in payments_by_acct[r['account_id']]:
            score = 0; reasons = []
            pAmt = float(p['amount']) if p['amount'] is not None else None
            if fee and pAmt is not None and abs(pAmt - fee) < 1:
                score += 5; reasons.append('amount matches')
            elif fee and pAmt is not None and fee > 0 and abs(pAmt - fee) / fee < 0.05:
                score += 3; reasons.append('amount close')
            pDate = p['paid_date'] or p['issue_date'] or p['due_date']
            if signed and pDate:
                days = abs((pDate - signed.date()).days) if hasattr(pDate, 'days') or hasattr(pDate, 'isoformat') else None
                if days is not None:
                    if days < 7: score += 3; reasons.append('same week')
                    elif days < 30: score += 1; reasons.append('same month')
            if score > 0 and (not best or score > best['score']):
                best = {'p': p, 'score': score, 'reasons': reasons}
        if best and best['score'] >= 3:
            invoice_label = best['p']['invoice_number'] or '(no INV number)'
            invoice_note = ', '.join(best['reasons'])
        else:
            invoice_note = 'no matching payment yet — link later when money arrives'
    elif not r['account_id']:
        invoice_note = '(no account, so no invoice to link)'
    else:
        invoice_note = 'no payments for this account yet'

    # Annual fee — show n/a for non-annual contract types; flag corrupted long values
    annual_fee_raw = (r['annual_fee'] or '').strip() if r['annual_fee'] is not None else ''
    annual_fee_display = annual_fee_raw
    annual_fee_issue = None
    if ct in ('formation','onboarding','tax_return','itin','custom') and not annual_fee_raw:
        annual_fee_display = 'n/a'
    if annual_fee_raw and len(annual_fee_raw.replace('.', '').replace(',', '')) > 6:
        # e.g. '200010001000' — string-concat bug
        annual_fee_issue = f'corrupted value ({annual_fee_raw}) — string-concat bug, real fee must be confirmed'

    # Possible mis-labeled contract type: renewal but company formed < 90 days before signing
    type_issue = None
    if ct == 'renewal' and r['account_formation_date'] and r['signed_at']:
        days = (r['signed_at'].date() - r['account_formation_date']).days
        if days < 90:
            type_issue = f'offer says renewal but company was formed only {days} days before signing — likely formation/onboarding'

    # Row classification
    notes_for_antonio = []
    if annual_fee_issue: notes_for_antonio.append(annual_fee_issue)
    if type_issue: notes_for_antonio.append(type_issue)
    if r['offer_token'] in DUPLICATE_TOKENS:
        classification = 'DUPLICATE — area 6'
    elif 'test-demo' in (r['offer_token'] or '').lower() or (r['client_email'] or '').lower() in ('sdfassf','ghfhfhj'):
        classification = 'SKIP — test data'
    elif contact_conf == 'LOW' or notes_for_antonio:
        classification = 'NEEDS REVIEW'
    else:
        classification = 'AUTO-APPLY'

    signed_date = r['signed_at'].date().isoformat() if r['signed_at'] else ''
    rows_out.append({
        'Action': classification,
        'Client Name': r['client_name'] or '',
        'Email': r['client_email'] or '',
        'Company': company,
        'Contract Type': human_contract_type(r['contract_type']),
        'Signed On': signed_date,
        'Annual Fee': annual_fee_display,
        'Proposed Contact': contact_name or '(none)',
        'Contact Note': contact_note,
        'Proposed Payment Type': payment_type,
        'Proposed Invoice': invoice_label or '(none)',
        'Invoice Note': invoice_note,
        'Services in Offer': len(r['services']) if isinstance(r['services'], list) else 0,
        'Offer Token': r['offer_token'] or '',
        'Review Notes': ' | '.join(notes_for_antonio) if notes_for_antonio else '',
    })

# Split into Main (unique tokens) and Duplicates (area 6 material)
duplicates_sheet = [r for r in rows_out if r['Action'] == 'DUPLICATE — area 6']
main_sheet = [r for r in rows_out if r['Action'] != 'DUPLICATE — area 6']

# Sort so NEEDS REVIEW is at top, then AUTO-APPLY, then SKIP
priority = {'NEEDS REVIEW': 0, 'AUTO-APPLY': 1, 'SKIP — test data': 2}
main_sheet.sort(key=lambda r: (priority.get(r['Action'], 99), r['Company'], r['Client Name']))
duplicates_sheet.sort(key=lambda r: (r['Offer Token'], r['Signed On']))
rows_out = main_sheet

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = 'Contracts Backfill Proposal'

headers = list(rows_out[0].keys())
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', start_color='305496')
review_fill = PatternFill('solid', start_color='FFF2CC')
apply_fill = PatternFill('solid', start_color='E2EFDA')
skip_fill = PatternFill('solid', start_color='F2F2F2')
thin = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for i, r in enumerate(rows_out, start=2):
    fill = review_fill if r['Action']=='NEEDS REVIEW' else apply_fill if r['Action']=='AUTO-APPLY' else skip_fill
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=i, column=c, value=r[h])
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Column widths
widths = {
    'Action': 18, 'Client Name': 28, 'Email': 30, 'Company': 34,
    'Contract Type': 22, 'Signed On': 12, 'Annual Fee': 14,
    'Proposed Contact': 28, 'Contact Note': 34,
    'Proposed Payment Type': 20, 'Proposed Invoice': 24, 'Invoice Note': 40,
    'Services in Offer': 10, 'Offer Token': 36,
    'Review Notes': 55,
}
for col_idx, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 18)
ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'

# Second sheet: DUPLICATES (for area 6 cleanup)
if duplicates_sheet:
    ws_dup = wb.create_sheet('Duplicates (area 6)')
    dup_headers = list(duplicates_sheet[0].keys())
    for c, h in enumerate(dup_headers, 1):
        cell = ws_dup.cell(row=1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    dup_fill = PatternFill('solid', start_color='FCE4D6')  # light orange
    for i, r in enumerate(duplicates_sheet, start=2):
        for c, h in enumerate(dup_headers, 1):
            cell = ws_dup.cell(row=i, column=c, value=r[h])
            cell.fill = dup_fill; cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for col_idx, h in enumerate(dup_headers, 1):
        ws_dup.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 18)
    ws_dup.row_dimensions[1].height = 30
    ws_dup.freeze_panes = 'A2'

# Third sheet: counts / summary
ws2 = wb.create_sheet('Summary')
total_all = len(rows_out) + len(duplicates_sheet)
ws2['A1'] = 'Total contracts in sandbox'; ws2['B1'] = total_all
ws2['A2'] = 'Main sheet — AUTO-APPLY (write as proposed)'; ws2['B2'] = sum(1 for r in rows_out if r['Action']=='AUTO-APPLY')
ws2['A3'] = 'Main sheet — NEEDS REVIEW'; ws2['B3'] = sum(1 for r in rows_out if r['Action']=='NEEDS REVIEW')
ws2['A4'] = 'Main sheet — SKIP (test data)'; ws2['B4'] = sum(1 for r in rows_out if r['Action']=='SKIP — test data')
ws2['A5'] = 'Duplicates sheet (area 6 cleanup)'; ws2['B5'] = len(duplicates_sheet)
ws2['A7'] = 'By Payment Type (main sheet)'
pt_counts = {}
for r in rows_out:
    pt_counts[r['Proposed Payment Type']] = pt_counts.get(r['Proposed Payment Type'], 0) + 1
r_idx = 8
for k, v in sorted(pt_counts.items(), key=lambda x: -x[1]):
    ws2[f'A{r_idx}'] = k; ws2[f'B{r_idx}'] = v; r_idx += 1
for col in ('A', 'B'):
    ws2.column_dimensions[col].width = 44 if col == 'A' else 10
for row in (1,2,3,4,5,7):
    ws2[f'A{row}'].font = Font(bold=True)

out = '/tmp/contracts-backfill-proposal.xlsx'
wb.save(out)
print(f'wrote {out}')
print(f'rows: {len(rows_out)}')
print(f"action breakdown: AUTO-APPLY={sum(1 for r in rows_out if r['Action']=='AUTO-APPLY')}, "
      f"NEEDS REVIEW={sum(1 for r in rows_out if r['Action']=='NEEDS REVIEW')}, "
      f"SKIP={sum(1 for r in rows_out if r['Action']=='SKIP — test data')}")
cur.close()
conn.close()
